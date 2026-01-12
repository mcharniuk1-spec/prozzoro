import os
import re
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional, List, Tuple

import pandas as pd
from openpyxl import load_workbook


DATA_DIR = Path("data")
OUTPUTS_DIR = DATA_DIR / "outputs"
LOGS_DIR = DATA_DIR / "logs"

EXPORT_RUN_ID = os.getenv("EXPORT_RUN_ID", "latest")  # "latest" or explicit run folder name
XLSX_OUT_DIR = Path(os.getenv("XLSX_OUT_DIR", str(OUTPUTS_DIR / "_xlsx_snapshots")))
MAX_ROWS_PER_SHEET = int(os.getenv("MAX_ROWS_PER_SHEET", "1000000"))  # <= 1,048,576


RUN_ID_RE = re.compile(r"^\d{8}_\d{6}$")   # your run_id format: YYYYMMDD_HHMMSS
MONTH_RE = re.compile(r"^\d{4}-\d{2}$")


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def setup_logger() -> logging.Logger:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    log_path = LOGS_DIR / f"export_{ts}.txt"

    logger = logging.getLogger("export_xlsx")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    sh = logging.StreamHandler()
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    logger.info("Log file: %s", log_path.as_posix())
    return logger


def find_latest_run_dir(logger: logging.Logger) -> Optional[Path]:
    if not OUTPUTS_DIR.exists():
        logger.warning("No outputs dir: %s", OUTPUTS_DIR.as_posix())
        return None

    candidates = []
    for p in OUTPUTS_DIR.iterdir():
        if p.is_dir() and RUN_ID_RE.match(p.name):
            candidates.append(p)

    if not candidates:
        logger.warning("No run folders found in %s", OUTPUTS_DIR.as_posix())
        return None

    # pick by modification time
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


def get_run_dir(logger: logging.Logger) -> Optional[Path]:
    if EXPORT_RUN_ID != "latest":
        p = OUTPUTS_DIR / EXPORT_RUN_ID
        if p.exists() and p.is_dir():
            return p
        logger.error("EXPORT_RUN_ID=%s not found under %s", EXPORT_RUN_ID, OUTPUTS_DIR.as_posix())
        return None

    return find_latest_run_dir(logger)


def read_csv_or_gz(path_csv: Path) -> Optional[pd.DataFrame]:
    if path_csv.exists():
        return pd.read_csv(path_csv)
    gz = Path(str(path_csv) + ".gz")
    if gz.exists():
        return pd.read_csv(gz, compression="gzip")
    return None


def sheet_chunks(df: pd.DataFrame, max_rows: int) -> List[Tuple[str, pd.DataFrame]]:
    if len(df) <= max_rows:
        return [("data", df)]
    chunks = []
    n = len(df)
    k = 1
    for start in range(0, n, max_rows):
        end = min(start + max_rows, n)
        chunks.append((f"part_{k}", df.iloc[start:end].copy()))
        k += 1
    return chunks


def finalize_workbook(path: Path, freeze_cell: str = "A2"):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = freeze_cell
        # AutoFilter on first row
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def export_month(logger: logging.Logger, run_dir: Path, month_dir: Path, out_dir: Path):
    lots_csv = month_dir / "dairy_lots.csv"
    items_csv = month_dir / "dairy_items.csv"

    lots_df = read_csv_or_gz(lots_csv)
    items_df = read_csv_or_gz(items_csv)

    if lots_df is None and items_df is None:
        logger.info("Skip %s (no lots/items csv)", month_dir.name)
        return

    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / f"dairy_{month_dir.name}.xlsx"

    logger.info("Exporting month=%s -> %s", month_dir.name, xlsx_path.as_posix())

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        if lots_df is not None:
            chunks = sheet_chunks(lots_df, MAX_ROWS_PER_SHEET)
            if len(chunks) == 1:
                chunks[0][1].to_excel(writer, sheet_name="lots", index=False)
            else:
                for i, (_, cdf) in enumerate(chunks, start=1):
                    cdf.to_excel(writer, sheet_name=f"lots_{i}", index=False)

        if items_df is not None:
            chunks = sheet_chunks(items_df, MAX_ROWS_PER_SHEET)
            if len(chunks) == 1:
                chunks[0][1].to_excel(writer, sheet_name="items", index=False)
            else:
                for i, (_, cdf) in enumerate(chunks, start=1):
                    cdf.to_excel(writer, sheet_name=f"items_{i}", index=False)

    finalize_workbook(xlsx_path)
    logger.info("Saved XLSX: %s", xlsx_path.as_posix())


def main():
    logger = setup_logger()
    logger.info("Started at %s", utc_now_iso())

    run_dir = get_run_dir(logger)
    if run_dir is None:
        logger.error("No run dir to export. Make sure fetch produced data/outputs/<run_id>/YYYY-MM/* files.")
        raise SystemExit(1)

    logger.info("Using run dir: %s", run_dir.as_posix())

    month_dirs = [p for p in run_dir.iterdir() if p.is_dir() and MONTH_RE.match(p.name)]
    if not month_dirs:
        logger.warning("No month dirs found in run dir: %s", run_dir.as_posix())
        raise SystemExit(0)

    month_dirs.sort(key=lambda x: x.name)

    out_base = XLSX_OUT_DIR / run_dir.name
    for md in month_dirs:
        export_month(logger, run_dir, md, out_base)

    logger.info("DONE. XLSX snapshots: %s", out_base.as_posix())


if __name__ == "__main__":
    main()
