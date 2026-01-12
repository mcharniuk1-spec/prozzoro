import os
import sys
import json
import csv
import time
import random
import pathlib
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import requests
from dateutil import parser as dtparser

from openpyxl import Workbook, load_workbook

BASE_URL = os.getenv("PROZORRO_BASE_URL", "https://public-api.prozorro.gov.ua/api/2.5").rstrip("/")
LIMIT = int(os.getenv("PROZORRO_LIMIT", "100"))
REQUEST_SLEEP_SECONDS = float(os.getenv("REQUEST_SLEEP_SECONDS", "0.15"))

MAX_RUNTIME_MINUTES = int(os.getenv("MAX_RUNTIME_MINUTES", "30"))  # важливо: коротко
SAVE_EVERY_SECONDS = int(os.getenv("SAVE_EVERY_SECONDS", "120"))   # CSV flush
XLSX_EVERY_SECONDS = int(os.getenv("XLSX_EVERY_SECONDS", "180"))   # XLSX update

# STRICT CPV list (ваші категорії)
ALLOWED_CPV_CODES = [
    "15500000",  # Dairy products
    "15510000",  # Milk and cream
    "15511000",  # Milk
    "15511100",  # Pasteurised milk
    "15511210",  # UHT milk
    "15512000",  # Cream
    "15530000",  # Butter
    "15540000",  # Cheese products
    "15550000",  # Assorted dairy products
]
ALLOWED_PREFIXES = [c.rstrip("0") for c in ALLOWED_CPV_CODES]  # e.g. 1551, 15511, 1551121...

ROOT = pathlib.Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
OUT_DIR = DATA_DIR / "outputs"
LOG_DIR = DATA_DIR / "logs"
CP_DIR = DATA_DIR / "checkpoints"

CURRENT_DIR = OUT_DIR / "current"
BY_MONTH_DIR = OUT_DIR / "by_month"

CHECKPOINT_PATH = CP_DIR / "checkpoint.json"

def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def iso(ts: datetime) -> str:
    return ts.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

def ensure_dirs():
    for p in [OUT_DIR, LOG_DIR, CP_DIR, CURRENT_DIR, BY_MONTH_DIR]:
        p.mkdir(parents=True, exist_ok=True)

def open_log_file() -> pathlib.Path:
    run_id = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    log_path = LOG_DIR / f"run_{run_id}.txt"
    return log_path

def log_line(fp, level: str, msg: str):
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{ts} | {level:<5} | {msg}\n"
    fp.write(line)
    fp.flush()

def load_checkpoint() -> Dict[str, Any]:
    if CHECKPOINT_PATH.exists():
        try:
            return json.loads(CHECKPOINT_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}

def save_checkpoint(obj: Dict[str, Any]):
    CHECKPOINT_PATH.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def normalize_cpv(raw: Optional[str]) -> str:
    if not raw:
        return ""
    digits = "".join(ch for ch in raw if ch.isdigit())
    return digits[:8]  # 15511210-8 -> 15511210

def cpv_is_dairy(cpv8: str) -> bool:
    if not cpv8:
        return False
    core = cpv8.rstrip("0")
    # allow by prefixes derived from your list
    return any(core.startswith(pref) for pref in ALLOWED_PREFIXES)

def safe_get(session: requests.Session, url: str, params: Optional[dict] = None, max_tries: int = 6) -> Dict[str, Any]:
    last_err = None
    for attempt in range(1, max_tries + 1):
        try:
            r = session.get(url, params=params, timeout=60)
            if r.status_code in (429, 500, 502, 503, 504):
                # backoff
                sleep_s = min(10.0, 0.5 * (2 ** (attempt - 1)) + random.random())
                time.sleep(sleep_s)
                continue
            r.raise_for_status()
            return r.json()
        except Exception as e:
            last_err = e
            time.sleep(min(10.0, 0.5 * (2 ** (attempt - 1)) + random.random()))
    raise RuntimeError(f"GET failed after retries: {url} | last={last_err}")

def list_tenders(session: requests.Session, offset: Optional[str]) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    url = f"{BASE_URL}/tenders"
    params = {"limit": LIMIT}
    if offset:
        params["offset"] = offset
    data = safe_get(session, url, params=params)
    items = data.get("data", []) or []
    next_offset = None
    np = data.get("next_page") or {}
    next_offset = np.get("offset")
    return items, next_offset

def fetch_tender_detail(session: requests.Session, tender_id: str) -> Dict[str, Any]:
    url = f"{BASE_URL}/tenders/{tender_id}"
    data = safe_get(session, url)
    return data.get("data") or {}

def month_key(dt_str: Optional[str]) -> str:
    if not dt_str:
        return "unknown"
    try:
        dt = dtparser.parse(dt_str)
        return dt.strftime("%Y-%m")
    except Exception:
        return "unknown"

def csv_path(kind: str, month: str) -> pathlib.Path:
    return BY_MONTH_DIR / f"{kind}_{month}.csv"

def ensure_csv_header(path: pathlib.Path, fieldnames: List[str]):
    if path.exists() and path.stat().st_size > 0:
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()

class XlsxAppender:
    def __init__(self, path: pathlib.Path):
        self.path = path
        self._wb = None

    def _init_if_needed(self):
        if self.path.exists():
            self._wb = load_workbook(self.path)
        else:
            self._wb = Workbook()
            # remove default sheet
            default = self._wb.active
            self._wb.remove(default)
            for name, headers in [
                ("tenders", ["tender_id","dateModified","dateCreated","status","buyer_name","buyer_id","procurementMethodType","title","value_amount","value_currency"]),
                ("lots",    ["tender_id","lot_id","lot_title","lot_status","lot_value_amount","lot_value_currency","minimalStep_amount","minimalStep_currency"]),
                ("items",   ["tender_id","lot_id","item_id","description","cpv","cpv_description","quantity","unit_name","unit_code","delivery_start","delivery_end","delivery_region","delivery_locality"]),
            ]:
                ws = self._wb.create_sheet(title=name)
                ws.append(headers)
            self._wb.save(self.path)

    def append_rows(self, sheet: str, rows: List[List[Any]]):
        if not rows:
            return
        self._init_if_needed()
        ws = self._wb[sheet]
        for r in rows:
            ws.append(r)

    def save(self):
        if self._wb:
            self._wb.save(self.path)

def main():
    ensure_dirs()
    log_path = open_log_file()

    cp = load_checkpoint()
    offset = cp.get("next_offset")  # server-provided offset
    started = now_utc()

    xlsx_path = CURRENT_DIR / "dairy_current.xlsx"
    xlsx = XlsxAppender(xlsx_path)

    # buffers
    tender_rows_csv: Dict[str, List[Dict[str, Any]]] = {}
    lot_rows_csv: Dict[str, List[Dict[str, Any]]] = {}
    item_rows_csv: Dict[str, List[Dict[str, Any]]] = {}

    # xlsx rows (list-of-lists)
    tender_rows_x: List[List[Any]] = []
    lot_rows_x: List[List[Any]] = []
    item_rows_x: List[List[Any]] = []

    last_save = time.time()
    last_xlsx = time.time()

    stats = {
        "pages": 0,
        "tenders_scanned": 0,
        "details_fetched": 0,
        "dairy_tenders": 0,
        "dairy_items": 0,
        "started_at": iso(started),
        "base_url": BASE_URL,
        "limit": LIMIT,
    }

    with log_path.open("w", encoding="utf-8") as lf, requests.Session() as session:
        log_line(lf, "INFO", f"Log file: {log_path}")
        log_line(lf, "INFO", f"BASE_URL={BASE_URL}")
        log_line(lf, "INFO", f"ALLOWED_PREFIXES={ALLOWED_PREFIXES}")
        log_line(lf, "INFO", f"Start offset={offset}")

        while True:
            elapsed = (now_utc() - started).total_seconds()
            if elapsed >= max(60, MAX_RUNTIME_MINUTES * 60 - 60):
                log_line(lf, "INFO", "Time budget reached, stopping gracefully...")
                break

            items, next_offset = list_tenders(session, offset)
            stats["pages"] += 1
            if not items:
                log_line(lf, "INFO", "No more tenders in list (empty page).")
                offset = next_offset
                break

            for rec in items:
                stats["tenders_scanned"] += 1
                tender_id = rec.get("id")
                if not tender_id:
                    continue

                # FULL mode: fetch details for every tender to avoid missing dairy
                detail = fetch_tender_detail(session, tender_id)
                stats["details_fetched"] += 1

                date_modified = detail.get("dateModified")
                date_created = detail.get("dateCreated")
                mkey = month_key(date_modified or date_created)

                title = detail.get("title") or ""
                status = detail.get("status") or ""

                buyer = (detail.get("procuringEntity") or {})
                buyer_name = buyer.get("name") or ""
                buyer_ident = (buyer.get("identifier") or {})
                buyer_id = buyer_ident.get("id") or ""

                pmtype = detail.get("procurementMethodType") or ""

                value = (detail.get("value") or {})
                value_amount = value.get("amount")
                value_currency = value.get("currency")

                # filter items строго по CPV
                raw_items = detail.get("items") or []
                dairy_items = []
                for it in raw_items:
                    cls = (it.get("classification") or {})
                    cpv8 = normalize_cpv(cls.get("id"))
                    if not cpv_is_dairy(cpv8):
                        continue
                    dairy_items.append(it)

                if not dairy_items:
                    # не молочний тендер -> пропускаємо
                    offset = next_offset
                    continue

                stats["dairy_tenders"] += 1

                tender_row = {
                    "tender_id": tender_id,
                    "dateModified": date_modified,
                    "dateCreated": date_created,
                    "status": status,
                    "buyer_name": buyer_name,
                    "buyer_id": buyer_id,
                    "procurementMethodType": pmtype,
                    "title": title,
                    "value_amount": value_amount,
                    "value_currency": value_currency,
                }
                tender_rows_csv.setdefault(mkey, []).append(tender_row)
                tender_rows_x.append([
                    tender_id, date_modified, date_created, status, buyer_name, buyer_id, pmtype, title, value_amount, value_currency
                ])

                # lots (якщо є)
                lots = detail.get("lots") or []
                lot_map = {l.get("id"): l for l in lots if l.get("id")}

                for it in dairy_items:
                    item_id = it.get("id") or ""
                    lot_id = it.get("relatedLot") or ""
                    desc = it.get("description") or ""
                    cls = (it.get("classification") or {})
                    cpv8 = normalize_cpv(cls.get("id"))
                    cpv_desc = cls.get("description") or ""

                    qty = it.get("quantity")
                    unit = (it.get("unit") or {})
                    unit_name = unit.get("name") or ""
                    unit_code = unit.get("code") or ""

                    ddate = (it.get("deliveryDate") or {})
                    d_start = ddate.get("startDate")
                    d_end = ddate.get("endDate")

                    addr = (it.get("deliveryAddress") or {})
                    region = addr.get("region") or ""
                    locality = addr.get("locality") or ""

                    item_row = {
                        "tender_id": tender_id,
                        "lot_id": lot_id,
                        "item_id": item_id,
                        "description": desc,
                        "cpv": cpv8,
                        "cpv_description": cpv_desc,
                        "quantity": qty,
                        "unit_name": unit_name,
                        "unit_code": unit_code,
                        "delivery_start": d_start,
                        "delivery_end": d_end,
                        "delivery_region": region,
                        "delivery_locality": locality,
                    }
                    item_rows_csv.setdefault(mkey, []).append(item_row)
                    item_rows_x.append([
                        tender_id, lot_id, item_id, desc, cpv8, cpv_desc, qty, unit_name, unit_code, d_start, d_end, region, locality
                    ])
                    stats["dairy_items"] += 1

                    if lot_id and lot_id in lot_map:
                        l = lot_map[lot_id]
                        lval = (l.get("value") or {})
                        ms = (l.get("minimalStep") or {})
                        lot_row = {
                            "tender_id": tender_id,
                            "lot_id": lot_id,
                            "lot_title": l.get("title"),
                            "lot_status": l.get("status"),
                            "lot_value_amount": lval.get("amount"),
                            "lot_value_currency": lval.get("currency"),
                            "minimalStep_amount": ms.get("amount"),
                            "minimalStep_currency": ms.get("currency"),
                        }
                        lot_rows_csv.setdefault(mkey, []).append(lot_row)
                        lot_rows_x.append([
                            tender_id, lot_id, l.get("title"), l.get("status"), lval.get("amount"), lval.get("currency"), ms.get("amount"), ms.get("currency")
                        ])

                # pacing
                time.sleep(REQUEST_SLEEP_SECONDS)

                # periodic CSV flush
                if time.time() - last_save >= SAVE_EVERY_SECONDS:
                    flush_csv(lf, tender_rows_csv, lot_rows_csv, item_rows_csv)
                    tender_rows_csv.clear()
                    lot_rows_csv.clear()
                    item_rows_csv.clear()
                    last_save = time.time()

                # periodic XLSX update
                if time.time() - last_xlsx >= XLSX_EVERY_SECONDS:
                    flush_xlsx(lf, xlsx, tender_rows_x, lot_rows_x, item_rows_x)
                    tender_rows_x.clear()
                    lot_rows_x.clear()
                    item_rows_x.clear()
                    last_xlsx = time.time()

            offset = next_offset

            if stats["pages"] % 50 == 0:
                log_line(lf, "INFO",
                         f"Progress | pages={stats['pages']} | scanned={stats['tenders_scanned']} | details={stats['details_fetched']} "
                         f"| dairy_tenders={stats['dairy_tenders']} | dairy_items={stats['dairy_items']} | offset={offset}")

        # final flush
        flush_csv(lf, tender_rows_csv, lot_rows_csv, item_rows_csv)
        flush_xlsx(lf, xlsx, tender_rows_x, lot_rows_x, item_rows_x)

        cp_out = {
            "next_offset": offset,
            "updated_at": iso(now_utc()),
            "stats": stats,
        }
        save_checkpoint(cp_out)
        log_line(lf, "INFO", f"Saved checkpoint: {CHECKPOINT_PATH}")
        log_line(lf, "INFO", f"Saved XLSX current: {xlsx_path}")
        log_line(lf, "INFO", f"Done. stats={json.dumps(stats, ensure_ascii=False)}")

def flush_csv(lf, tender_rows_csv, lot_rows_csv, item_rows_csv):
    # tenders
    tender_fields = ["tender_id","dateModified","dateCreated","status","buyer_name","buyer_id","procurementMethodType","title","value_amount","value_currency"]
    lot_fields = ["tender_id","lot_id","lot_title","lot_status","lot_value_amount","lot_value_currency","minimalStep_amount","minimalStep_currency"]
    item_fields = ["tender_id","lot_id","item_id","description","cpv","cpv_description","quantity","unit_name","unit_code","delivery_start","delivery_end","delivery_region","delivery_locality"]

    for m, rows in tender_rows_csv.items():
        p = csv_path("tenders", m)
        ensure_csv_header(p, tender_fields)
        with p.open("a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=tender_fields)
            w.writerows(rows)

    for m, rows in lot_rows_csv.items():
        p = csv_path("lots", m)
        ensure_csv_header(p, lot_fields)
        with p.open("a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=lot_fields)
            w.writerows(rows)

    for m, rows in item_rows_csv.items():
        p = csv_path("items", m)
        ensure_csv_header(p, item_fields)
        with p.open("a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=item_fields)
            w.writerows(rows)

    log_line(lf, "INFO", f"CSV flush: tenders_months={len(tender_rows_csv)} lots_months={len(lot_rows_csv)} items_months={len(item_rows_csv)}")

def flush_xlsx(lf, xlsx: XlsxAppender, tender_rows_x, lot_rows_x, item_rows_x):
    xlsx.append_rows("tenders", tender_rows_x)
    xlsx.append_rows("lots", lot_rows_x)
    xlsx.append_rows("items", item_rows_x)
    xlsx.save()
    log_line(lf, "INFO", f"XLSX update: tenders+{len(tender_rows_x)} lots+{len(lot_rows_x)} items+{len(item_rows_x)}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted", file=sys.stderr)
        sys.exit(130)
